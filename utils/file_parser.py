import pandas as pd
import pypdf
import io

import openpyxl

def parse_excel(file):
    """
    Parses an Excel file and returns a dictionary for each sheet containing values and formulas.
    """
    # 1. Read values using pandas
    file.seek(0)
    # Determine engine based on file type or let pandas decide, but we need to handle the openpyxl error specifically for .xls
    # We'll try to read with pandas first. Pandas usually auto-detects.
    # For .xls we need 'xlrd', for .xlsx 'openpyxl'.
    
    try:
        xls = pd.ExcelFile(file)
    except Exception:
        # Fallback if auto-detection fails, though it shouldn't with valid files
        file.seek(0)
        xls = pd.ExcelFile(file, engine='openpyxl')

    sheet_values = {}
    for sheet_name in xls.sheet_names:
        sheet_values[sheet_name] = pd.read_excel(xls, sheet_name=sheet_name, header=None)
        
    # 2. Read formulas using openpyxl (ONLY for .xlsx)
    sheet_formulas = {}
    try:
        file.seek(0)
        wb = openpyxl.load_workbook(file, data_only=False)
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            formulas = {}
            for row in ws.iter_rows():
                for cell in row:
                    if isinstance(cell.value, str) and cell.value.startswith('='):
                        formulas[cell.coordinate] = cell.value
            sheet_formulas[sheet_name] = formulas
    except Exception as e:
        # This likely means it's an .xls file or not supported by openpyxl
        # We just skip formula extraction
        print(f"Could not extract formulas (likely .xls file): {e}")
        pass
        
    # 3. Combine results
    sheets_data = {}
    for sheet_name in sheet_values.keys():
        sheets_data[sheet_name] = {
            "values": sheet_values[sheet_name],
            "formulas": sheet_formulas.get(sheet_name, {})
        }
        
    return sheets_data

def parse_pdf(file):
    """
    Extracts text from a PDF file.
    """
    pdf_reader = pypdf.PdfReader(file)
    text = ""
    for page in pdf_reader.pages:
        text += page.extract_text() + "\n"
    return text
